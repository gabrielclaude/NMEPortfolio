"""
Seed Resource Management tables from RMSimplifiedV01082025SPG.xlsx
Creates tables and populates:
  rm_fte_matrix, rm_study, rm_study_segment,
  rm_monthly_fte, rm_personnel, rm_staff_assignment
"""
import os, datetime
import openpyxl
import psycopg2
from psycopg2.extras import execute_values

XLSX = "/Users/claudegabriel/Documents/2026/PSEC/RMSimplifiedV01082025SPG.xlsx"
DB   = os.environ.get("DATABASE_URL", "postgresql://claudegabriel@localhost:5433/nme_portfolio")

conn = psycopg2.connect(DB)
cur  = conn.cursor()

# ─── Create Tables ───────────────────────────────────────────────────────────
cur.execute("""
DROP TABLE IF EXISTS rm_staff_assignment CASCADE;
DROP TABLE IF EXISTS rm_monthly_fte CASCADE;
DROP TABLE IF EXISTS rm_study_segment CASCADE;
DROP TABLE IF EXISTS rm_personnel CASCADE;
DROP TABLE IF EXISTS rm_study CASCADE;
DROP TABLE IF EXISTS rm_fte_matrix CASCADE;
""")

cur.execute("""
CREATE TABLE rm_fte_matrix (
  id          SERIAL PRIMARY KEY,
  complexity  TEXT NOT NULL,
  role        TEXT NOT NULL,
  phase       INT  NOT NULL,
  activity    TEXT NOT NULL,
  fte_per_month FLOAT NOT NULL,
  UNIQUE(complexity, role, phase, activity)
);

CREATE TABLE rm_study (
  id         TEXT PRIMARY KEY,
  phase      INT  NOT NULL DEFAULT 1,
  status     TEXT NOT NULL DEFAULT 'Active',
  complexity TEXT NOT NULL DEFAULT 'Low',
  nme_id     TEXT REFERENCES "NME"(id)
);

CREATE TABLE rm_study_segment (
  id           SERIAL PRIMARY KEY,
  study_id     TEXT REFERENCES rm_study(id),
  activity     TEXT NOT NULL,
  start_date   DATE NOT NULL,
  end_date     DATE NOT NULL,
  complexity   TEXT NOT NULL,
  role         TEXT NOT NULL,
  phase        INT  NOT NULL,
  days         INT  NOT NULL,
  fte_per_month FLOAT NOT NULL,
  UNIQUE(study_id, activity, role)
);

CREATE TABLE rm_monthly_fte (
  id          SERIAL PRIMARY KEY,
  segment_id  INT REFERENCES rm_study_segment(id) ON DELETE CASCADE,
  month_date  DATE NOT NULL,
  fte_value   FLOAT NOT NULL DEFAULT 0,
  UNIQUE(segment_id, month_date)
);

CREATE TABLE rm_personnel (
  id               SERIAL PRIMARY KEY,
  name             TEXT NOT NULL UNIQUE,
  total_allocation FLOAT NOT NULL DEFAULT 0,
  adjustment       FLOAT NOT NULL DEFAULT 0
);

CREATE TABLE rm_staff_assignment (
  id             SERIAL PRIMARY KEY,
  study_id       TEXT REFERENCES rm_study(id),
  personnel_id   INT  REFERENCES rm_personnel(id),
  role           TEXT NOT NULL,
  allocation_pct FLOAT NOT NULL,
  UNIQUE(study_id, personnel_id, role)
);
""")
conn.commit()
print("Tables created.")

# ─── Load Workbook ────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws_fte  = wb["Monthly FTE Dist"]
ws_alloc = wb["Allocation Tool"]

# ─── 1. FTE Attribute Matrix (rows 24–32, cols 1–5) ──────────────────────────
fte_matrix = []
for r in range(24, 33):   # rows 24–32 inclusive
    complexity = ws_fte.cell(r, 1).value
    role       = ws_fte.cell(r, 2).value
    phase      = ws_fte.cell(r, 3).value
    activity   = ws_fte.cell(r, 4).value
    fte        = ws_fte.cell(r, 5).value
    if complexity and role and phase and activity and fte is not None:
        fte_matrix.append((str(complexity).strip(), str(role).strip(),
                           int(phase), str(activity).strip(), float(fte)))

execute_values(cur,
    "INSERT INTO rm_fte_matrix (complexity,role,phase,activity,fte_per_month) VALUES %s",
    fte_matrix)
conn.commit()
print(f"FTE matrix: {len(fte_matrix)} rows inserted.")

# ─── 2. Month headers (row 29, cols 18–57) ───────────────────────────────────
# Col 18 = Dec-2023 (pre-period placeholder), col 19 = Jan-2024 start
month_cols = {}    # {col_index: date}
for c in range(18, 58):
    v = ws_fte.cell(29, c).value
    if isinstance(v, datetime.datetime):
        month_cols[c] = v.date().replace(day=1)   # normalize to first of month

print(f"Month columns found: {len(month_cols)} ({min(month_cols.values())} – {max(month_cols.values())})")

# Only keep Jan-2024 onward (exclude pre-period Dec-2023)
month_cols = {c: d for c, d in month_cols.items() if d >= datetime.date(2024, 1, 1)}
print(f"After filter (≥ Jan-2024): {len(month_cols)} months")

# ─── 3. Study Segments + Monthly FTE (rows 32–49, cols 6+) ───────────────────
# Collect all unique study IDs from segment rows first
segment_study_ids = set()
segment_rows = []
for r in range(32, 50):
    study_id = ws_fte.cell(r, 6).value
    if not study_id or str(study_id).strip() in ('-', ''):
        continue
    study_id = str(study_id).strip()
    phase_num = ws_fte.cell(r, 7).value
    activity  = ws_fte.cell(r, 8).value
    start_v   = ws_fte.cell(r, 9).value
    end_v     = ws_fte.cell(r, 10).value
    complexity= ws_fte.cell(r, 11).value
    role      = ws_fte.cell(r, 12).value
    days_v    = ws_fte.cell(r, 15).value
    fte_pm    = ws_fte.cell(r, 16).value

    if not all([phase_num, activity, start_v, end_v, complexity, role, days_v, fte_pm]):
        continue
    if str(activity).strip() in ('-', ''):
        continue

    start_date = start_v.date() if isinstance(start_v, datetime.datetime) else start_v
    end_date   = end_v.date()   if isinstance(end_v,   datetime.datetime) else end_v

    # Gather monthly FTE values
    monthly = {}
    for c, mo in month_cols.items():
        val = ws_fte.cell(r, c).value
        monthly[mo] = float(val) if val is not None else 0.0

    segment_study_ids.add(study_id)
    segment_rows.append({
        "study_id":     study_id,
        "phase":        int(phase_num),
        "activity":     str(activity).strip(),
        "start_date":   start_date,
        "end_date":     end_date,
        "complexity":   str(complexity).strip(),
        "role":         str(role).strip(),
        "days":         int(days_v),
        "fte_per_month":float(fte_pm),
        "monthly":      monthly,
    })

print(f"Study segments found: {len(segment_rows)} across studies: {sorted(segment_study_ids)}")

# ─── 4. Allocation Tool → Personnel + Assignments ────────────────────────────
# Header row 3: [3]=Study ID, [4]=CS name, [5]=CS%, [7]=MM name, [8]=MM%,
#               [10]=SMM name, [11]=SMM%, [13]=DTL name, [14]=DTL%,
#               [16]=Personnel, [17]=Total Allocation, [18]=Adjustment
ROLE_COLS = {
    "Clinical Scientist":        (4, 5),
    "Medical Monitor":           (7, 8),
    "Support Medical Monitor":  (10, 11),
    "Development Team Lead":    (13, 14),
}

alloc_study_ids = set()
assignments = []      # (study_id, person_name, role, pct)
personnel_map = {}    # name → total_allocation

# Read assignment rows (rows 4–10)
for r in range(4, 13):
    study_id = ws_alloc.cell(r, 3).value
    if study_id and str(study_id).strip() not in ('', '-'):
        sid = str(study_id).strip()
        alloc_study_ids.add(sid)
        for role_label, (name_col, pct_col) in ROLE_COLS.items():
            name = ws_alloc.cell(r, name_col).value
            pct  = ws_alloc.cell(r, pct_col).value
            if name and str(name).strip() not in ('', '-'):
                pct_val = float(pct) if pct is not None else 0.0
                assignments.append((sid, str(name).strip(), role_label, pct_val))

    # Personnel summary (cols 16–17 on every row 4–12)
    p_name  = ws_alloc.cell(r, 16).value
    p_alloc = ws_alloc.cell(r, 17).value
    p_adj   = ws_alloc.cell(r, 18).value
    if p_name and str(p_name).strip() not in ('', '-'):
        personnel_map[str(p_name).strip()] = (
            float(p_alloc) if p_alloc is not None else 0.0,
            float(p_adj)   if p_adj   is not None else 0.0,
        )

print(f"Allocation Tool studies: {sorted(alloc_study_ids)}")
print(f"Personnel: {list(personnel_map.keys())}")
print(f"Assignments: {len(assignments)}")

# ─── 5. Infer all studies (union of both sources) ────────────────────────────
all_study_ids = segment_study_ids | alloc_study_ids

# Derive phase from segment data where available
study_phase_map = {}
study_complexity_map = {}
for seg in segment_rows:
    study_phase_map[seg["study_id"]] = seg["phase"]
    study_complexity_map[seg["study_id"]] = seg["complexity"]

# Fetch available NME IDs to distribute studies across NMEs
cur.execute('SELECT id FROM "NME" ORDER BY code')
nme_ids = [row[0] for row in cur.fetchall()]
print(f"Found {len(nme_ids)} NMEs for study assignment")

# Insert studies with NME assignment (round-robin distribution)
study_rows = []
for i, sid in enumerate(sorted(all_study_ids)):
    phase      = study_phase_map.get(sid, 1)
    complexity = study_complexity_map.get(sid, "Low")
    # Assign NME in round-robin fashion if NMEs exist
    nme_id = nme_ids[i % len(nme_ids)] if nme_ids else None
    study_rows.append((sid, phase, "Active", complexity, nme_id))

execute_values(cur,
    "INSERT INTO rm_study (id, phase, status, complexity, nme_id) VALUES %s ON CONFLICT DO NOTHING",
    study_rows)
conn.commit()
print(f"Studies: {len(study_rows)} inserted with NME assignments.")

# ─── 6. Insert Personnel ─────────────────────────────────────────────────────
# Collect all unique person names from assignments too
all_person_names = set(personnel_map.keys())
for _, name, _, _ in assignments:
    all_person_names.add(name)

person_rows = []
for name in sorted(all_person_names):
    alloc, adj = personnel_map.get(name, (0.0, 0.0))
    person_rows.append((name, alloc, adj))

execute_values(cur,
    "INSERT INTO rm_personnel (name, total_allocation, adjustment) VALUES %s ON CONFLICT DO NOTHING",
    person_rows)
conn.commit()
print(f"Personnel: {len(person_rows)} inserted.")

# Build personnel name → id map
cur.execute("SELECT id, name FROM rm_personnel")
person_id_map = {row[1]: row[0] for row in cur.fetchall()}

# ─── 7. Insert Study Segments + Monthly FTE ──────────────────────────────────
for seg in segment_rows:
    cur.execute("""
        INSERT INTO rm_study_segment
          (study_id, activity, start_date, end_date, complexity, role, phase, days, fte_per_month)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (study_id, activity, role) DO UPDATE SET
          start_date=EXCLUDED.start_date, end_date=EXCLUDED.end_date,
          days=EXCLUDED.days, fte_per_month=EXCLUDED.fte_per_month
        RETURNING id
    """, (seg["study_id"], seg["activity"], seg["start_date"], seg["end_date"],
          seg["complexity"], seg["role"], seg["phase"], seg["days"], seg["fte_per_month"]))
    seg_id = cur.fetchone()[0]

    # Monthly FTE rows
    monthly_rows = [(seg_id, mo, val) for mo, val in seg["monthly"].items()]
    if monthly_rows:
        execute_values(cur,
            "INSERT INTO rm_monthly_fte (segment_id, month_date, fte_value) VALUES %s ON CONFLICT DO NOTHING",
            monthly_rows)

conn.commit()
print(f"Segments: {len(segment_rows)} inserted with monthly FTE.")

# ─── 8. Insert Staff Assignments ─────────────────────────────────────────────
assign_rows = []
for study_id, name, role, pct in assignments:
    pid = person_id_map.get(name)
    if pid:
        assign_rows.append((study_id, pid, role, pct))

execute_values(cur,
    """INSERT INTO rm_staff_assignment (study_id, personnel_id, role, allocation_pct)
       VALUES %s ON CONFLICT DO NOTHING""",
    assign_rows)
conn.commit()
print(f"Staff assignments: {len(assign_rows)} inserted.")

# ─── 9. Verification ─────────────────────────────────────────────────────────
cur.execute("SELECT COUNT(*) FROM rm_fte_matrix");        print(f"rm_fte_matrix rows: {cur.fetchone()[0]}")
cur.execute("SELECT COUNT(*) FROM rm_study");             print(f"rm_study rows:       {cur.fetchone()[0]}")
cur.execute("SELECT COUNT(*) FROM rm_study_segment");     print(f"rm_study_segment:    {cur.fetchone()[0]}")
cur.execute("SELECT COUNT(*) FROM rm_monthly_fte");       print(f"rm_monthly_fte:      {cur.fetchone()[0]}")
cur.execute("SELECT COUNT(*) FROM rm_personnel");         print(f"rm_personnel:        {cur.fetchone()[0]}")
cur.execute("SELECT COUNT(*) FROM rm_staff_assignment");  print(f"rm_staff_assignment: {cur.fetchone()[0]}")

cur.execute("""
  SELECT s.id, seg.activity, seg.role, seg.fte_per_month,
         SUM(m.fte_value) as total_fte
  FROM rm_study s
  JOIN rm_study_segment seg ON seg.study_id = s.id
  JOIN rm_monthly_fte m ON m.segment_id = seg.id
  GROUP BY s.id, seg.activity, seg.role, seg.fte_per_month
  ORDER BY s.id, seg.activity
""")
print("\nSegment summary:")
for row in cur.fetchall():
    print(f"  {row[0]:6s} | {row[1]:12s} | {row[2]:22s} | {row[3]:.2f} FTE/mo | {row[4]:.2f} total")

# ─── 10. Generate synthetic RM data for more NMEs ─────────────────────────────
import random
random.seed(42)  # For reproducibility

# First, add segments to existing studies that don't have any
cur.execute("""
    SELECT s.id, s.phase, s.complexity, s.nme_id
    FROM rm_study s
    LEFT JOIN rm_study_segment seg ON seg.study_id = s.id
    WHERE seg.id IS NULL AND s.nme_id IS NOT NULL
""")
studies_without_segments = cur.fetchall()
print(f"Adding segments to {len(studies_without_segments)} existing studies without segment data...")

roles = ["Clinical Scientist", "Medical Monitor", "Clinical RA"]
activities = ["Start Up", "Conduct", "Close Out"]
complexities = ["Low", "Medium", "High"]
fte_base = {"Low": 0.5, "Medium": 1.0, "High": 1.5}

for study_id, phase, complexity, nme_id in studies_without_segments:
    if not complexity:
        complexity = random.choice(complexities)

    # Create segments for each role
    for role in roles:
        start_month = random.randint(1, 12)
        start_date = datetime.date(2024, 1, 1) + datetime.timedelta(days=start_month * 30)

        for activity in activities:
            if activity == "Start Up":
                duration_months = random.randint(3, 6)
            elif activity == "Conduct":
                duration_months = random.randint(12, 18)
            else:
                duration_months = random.randint(3, 6)

            end_date = start_date + datetime.timedelta(days=duration_months * 30)
            days = (end_date - start_date).days
            fte_pm = fte_base.get(complexity, 1.0) * random.uniform(0.5, 1.5)

            cur.execute("""
                INSERT INTO rm_study_segment
                    (study_id, activity, start_date, end_date, complexity, role, phase, days, fte_per_month)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (study_id, activity, role) DO NOTHING
                RETURNING id
            """, (study_id, activity, start_date, end_date, complexity, role, phase or 1, days, round(fte_pm, 2)))

            result = cur.fetchone()
            if result:
                seg_id = result[0]
                monthly_rows = []
                current = datetime.date(start_date.year, start_date.month, 1)
                end_month = datetime.date(end_date.year, end_date.month, 1)

                while current <= end_month and current <= datetime.date(2026, 12, 1):
                    month_start = max(current, start_date)
                    next_month = datetime.date(current.year + (current.month // 12),
                                               ((current.month % 12) + 1), 1)
                    month_end = min(next_month - datetime.timedelta(days=1), end_date)

                    if month_end >= month_start:
                        days_in_month = (month_end - month_start).days + 1
                        total_days_in_month = (next_month - current).days
                        prorated_fte = fte_pm * (days_in_month / total_days_in_month)
                        monthly_rows.append((seg_id, current, round(prorated_fte, 4)))

                    current = next_month

                if monthly_rows:
                    execute_values(cur,
                        "INSERT INTO rm_monthly_fte (segment_id, month_date, fte_value) VALUES %s ON CONFLICT DO NOTHING",
                        monthly_rows)

            start_date = end_date

conn.commit()
print(f"Segments added to {len(studies_without_segments)} existing studies.")

# Now get NMEs that still don't have any studies with segments
cur.execute("""
    SELECT n.id, n.code FROM "NME" n
    WHERE n.id NOT IN (
        SELECT DISTINCT s.nme_id FROM rm_study s
        JOIN rm_study_segment seg ON seg.study_id = s.id
        WHERE s.nme_id IS NOT NULL
    )
    ORDER BY n.code
    LIMIT 12
""")
nmes_without_data = cur.fetchall()
print(f"Generating synthetic data for {len(nmes_without_data)} NMEs...")

roles = ["Clinical Scientist", "Medical Monitor", "Clinical RA"]
activities = ["Start Up", "Conduct", "Close Out"]
complexities = ["Low", "Medium", "High"]
fte_base = {"Low": 0.5, "Medium": 1.0, "High": 1.5}

for nme_id, nme_code in nmes_without_data:
    # Create 1-2 studies per NME
    num_studies = random.randint(1, 2)
    for study_num in range(num_studies):
        study_id = f"{nme_code}-S{study_num + 1}"
        phase = random.randint(1, 3)
        complexity = random.choice(complexities)

        # Insert study
        cur.execute("""
            INSERT INTO rm_study (id, phase, status, complexity, nme_id)
            VALUES (%s, %s, 'Active', %s, %s)
            ON CONFLICT DO NOTHING
        """, (study_id, phase, complexity, nme_id))

        # Create segments for each role
        for role in roles:
            # Randomize start dates within 2024-2025
            start_month = random.randint(1, 18)  # Jan 2024 to Jun 2025
            start_date = datetime.date(2024, 1, 1) + datetime.timedelta(days=start_month * 30)

            for activity in activities:
                # Duration varies by activity
                if activity == "Start Up":
                    duration_months = random.randint(3, 6)
                elif activity == "Conduct":
                    duration_months = random.randint(12, 24)
                else:  # Close Out
                    duration_months = random.randint(3, 6)

                end_date = start_date + datetime.timedelta(days=duration_months * 30)
                days = (end_date - start_date).days
                fte_pm = fte_base[complexity] * random.uniform(0.5, 1.5)

                # Insert segment
                cur.execute("""
                    INSERT INTO rm_study_segment
                        (study_id, activity, start_date, end_date, complexity, role, phase, days, fte_per_month)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (study_id, activity, role) DO NOTHING
                    RETURNING id
                """, (study_id, activity, start_date, end_date, complexity, role, phase, days, round(fte_pm, 2)))

                result = cur.fetchone()
                if result:
                    seg_id = result[0]

                    # Generate monthly FTE values
                    monthly_rows = []
                    current = datetime.date(start_date.year, start_date.month, 1)
                    end_month = datetime.date(end_date.year, end_date.month, 1)

                    while current <= end_month and current <= datetime.date(2026, 12, 1):
                        # Calculate prorated FTE for this month
                        month_start = max(current, start_date)
                        next_month = datetime.date(current.year + (current.month // 12),
                                                   ((current.month % 12) + 1), 1)
                        month_end = min(next_month - datetime.timedelta(days=1), end_date)

                        if month_end >= month_start:
                            days_in_month = (month_end - month_start).days + 1
                            total_days_in_month = (next_month - current).days
                            prorated_fte = fte_pm * (days_in_month / total_days_in_month)
                            monthly_rows.append((seg_id, current, round(prorated_fte, 4)))

                        current = next_month

                    if monthly_rows:
                        execute_values(cur,
                            "INSERT INTO rm_monthly_fte (segment_id, month_date, fte_value) VALUES %s ON CONFLICT DO NOTHING",
                            monthly_rows)

                # Move start date forward for next activity
                start_date = end_date

conn.commit()
print(f"Synthetic data generated for {len(nmes_without_data)} NMEs.")

# Also create some staff assignments for the synthetic studies
cur.execute("SELECT id FROM rm_personnel ORDER BY RANDOM() LIMIT 4")
personnel_ids = [row[0] for row in cur.fetchall()]

if personnel_ids:
    for nme_id, nme_code in nmes_without_data:
        cur.execute("SELECT id FROM rm_study WHERE nme_id = %s", (nme_id,))
        study_ids = [row[0] for row in cur.fetchall()]

        for study_id in study_ids:
            # Assign 1-2 staff members
            for role in random.sample(roles[:2], random.randint(1, 2)):
                pid = random.choice(personnel_ids)
                pct = random.uniform(0.3, 0.8)
                cur.execute("""
                    INSERT INTO rm_staff_assignment (study_id, personnel_id, role, allocation_pct)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT DO NOTHING
                """, (study_id, pid, role, round(pct, 2)))

conn.commit()
print("Staff assignments added for synthetic studies.")

# ─── 11. Create view for NME-level monthly FTE ────────────────────────────────
cur.execute("""
DROP VIEW IF EXISTS v_rm_monthly_by_nme;
CREATE OR REPLACE VIEW v_rm_monthly_by_nme AS
SELECT
  m.month_date,
  TO_CHAR(m.month_date, 'Mon-YY') AS month_label,
  s.nme_id,
  seg.role,
  ROUND(SUM(m.fte_value)::NUMERIC, 4) AS fte_demand
FROM rm_monthly_fte m
JOIN rm_study_segment seg ON seg.id = m.segment_id
JOIN rm_study s ON s.id = seg.study_id
WHERE s.nme_id IS NOT NULL
GROUP BY m.month_date, s.nme_id, seg.role
ORDER BY m.month_date, seg.role;
""")
conn.commit()
print("View v_rm_monthly_by_nme created.")

cur.close()
conn.close()
print("\nDone.")
